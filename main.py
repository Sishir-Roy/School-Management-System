from tkinter import *
from tkinter import ttk
from tkinter import Tk
from tkinter import messagebox
import openpyxl
import tkcalendar
from tkcalendar import DateEntry

class Student:
    def __init__(self, root):
        self.root = root
        self.root.geometry("1520x790+0+0")
        self.root.configure(bg="black")
        self.root.title("School Management System")
        title=Label(self.root, text="School Management System", font=("times new roman",30,"bold"), bg="white", fg="black")
        title.place(x=0, y=0, width=1520, height=50)

        self.var_id = StringVar()
        self.var_name = StringVar()
        self.var_gender = StringVar()
        self.var_email = StringVar()
        self.var_contact = StringVar()
        self.var_address = StringVar()
        self.var_dept = StringVar()
        self.var_session = StringVar()
        self.var_lt = StringVar()
        self.var_courses = StringVar()

        self.id_var = StringVar()
        self.name_var = StringVar()
        self.gender_var = StringVar()
        self.email_var = StringVar()
        self.contact_var = StringVar()
        self.address_var = StringVar()
        self.dept_var = StringVar()
        self.position_var = StringVar()
        self.joining_var = StringVar()
        self.session_var = StringVar()
        self.courses_var = StringVar()
        self.basic_var = StringVar()
        self.other_var = StringVar()
        self.net_var = StringVar()


        self.frame1 = LabelFrame(self.root, text="Student Info", font=("times new roman",16,"bold"), fg="red", bd=3, relief=RIDGE)
        self.frame1.place(x=10, y=60, width=740, height=300)

        self.frame2 = LabelFrame(self.root, text="Student Details", font=("times new roman",16,"bold"), fg="red", bd=3, relief=RIDGE)
        self.frame2.place(x=10, y=370, width=740, height=410)

        self.frame3 = LabelFrame(self.root, text="Teacher Info", font=("times new roman",16,"bold"), fg="red", bd=3, relief=RIDGE)
        self.frame3.place(x=770, y=60, width=740, height=355)

        self.frame4 = LabelFrame(self.root, text="Teacher Details", font=("times new roman",16,"bold"), fg="red", bd=3, relief=RIDGE)
        self.frame4.place(x=770, y=425, width=740, height=355)

        student_id=Label(self.frame1, text="Student ID :", font=("times new roman",14,"bold"), bg="white", bd=0, relief=RIDGE)
        student_id.grid(row=0, column=0, padx=5, pady=3, sticky=W)

        id_entry = ttk.Entry(self.frame1, textvariable=self.var_id, font=("times new roman", 14, "bold"), width=20)
        id_entry.grid(row=0, column=1, padx=3, pady=3, sticky=W)

        student_name = Label(self.frame1, text="Student Name :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        student_name.grid(row=1, column=0, padx=5, pady=3, sticky=W)

        name_entry = ttk.Entry(self.frame1, textvariable=self.var_name, font=("times new roman", 14, "bold"), width=20)
        name_entry.grid(row=1, column=1, padx=3, pady=3, sticky=W)

        gender = Label(self.frame1, text="Gender :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        gender.grid(row=2, column=0, padx=5, pady=3, sticky=W)

        gender_entry = ttk.Combobox(self.frame1, textvariable=self.var_gender, font=("times new roman", 14, "bold"), width=18, state="readonly")
        gender_entry["value"]=("Male","Female","Others")
        gender_entry.grid(row=2, column=1, padx=3, pady=3, sticky=W)

        email = Label(self.frame1, text="Email :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        email.grid(row=3, column=0, padx=5, pady=3, sticky=W)

        email_entry = ttk.Entry(self.frame1, textvariable=self.var_email, font=("times new roman", 14, "bold"), width=20)
        email_entry.grid(row=3, column=1, padx=3, pady=3, sticky=W)

        contact = Label(self.frame1, text="Contact No :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        contact.grid(row=4, column=0, padx=5, pady=3, sticky=W)

        contact_entry = ttk.Entry(self.frame1, textvariable=self.var_contact, font=("times new roman", 14, "bold"), width=20)
        contact_entry.grid(row=4, column=1, padx=3, pady=3, sticky=W)

        address = Label(self.frame1, text="Address :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        address.grid(row=0, column=1, padx=250, pady=3, sticky=W)

        address_entry = ttk.Entry(self.frame1, textvariable=self.var_address, font=("times new roman", 14, "bold"), width=20)
        address_entry.grid(row=0, column=1, padx=370, pady=3)

        department = Label(self.frame1, text="Department :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        department.grid(row=1, column=1, padx=250, pady=3, sticky=W)

        department_entry = ttk.Combobox(self.frame1, textvariable=self.var_dept, font=("times new roman", 14, "bold"), width=18, state="readonly")
        department_entry["value"] = ("CSE","EEE","ME","CE","IPE", "BME","ChE","WRE","NAME","URP","ARCHI")
        department_entry.grid(row=1, column=1, padx=370, pady=3)

        session = Label(self.frame1, text="Session :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        session.grid(row=2, column=1, padx=250, pady=3, sticky=W)

        session_entry = ttk.Combobox(self.frame1, textvariable=self.var_session, font=("times new roman", 14, "bold"), width=18, state="readonly")
        session_entry["value"] = ("2018-19", "2019-20", "2020-21","2021-22","2022-23","2023-24")
        session_entry.grid(row=2, column=1, padx=370, pady=3)

        level_term = Label(self.frame1, text="Level/Term :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        level_term.grid(row=3, column=1, padx=250, pady=3, sticky=W)

        level_term_entry = ttk.Combobox(self.frame1, textvariable=self.var_lt, font=("times new roman", 14, "bold"), width=18, state="readonly")
        level_term_entry["value"] = ("1-1", "1-2", "2-1", "2-2", "3-1", "3-2", "4-1", "4-2")
        level_term_entry.grid(row=3, column=1, padx=370, pady=3, sticky=W)

        courses = Label(self.frame1, text="Courses :", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        courses.grid(row=4, column=1, padx=250, pady=3, sticky=W)

        courses_entry = ttk.Entry(self.frame1, textvariable=self.var_courses, font=("times new roman", 14, "bold"), width=20)
        courses_entry.grid(row=4, column=1, padx=370, pady=3)


        button_frame = Frame(self.frame1, bd=3, relief=RIDGE)
        button_frame.place(x=5, y=205, width=725, height=45)

        add_button = Button(button_frame, text="Add", font=("verdana", 14, "bold"),width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.add_student)
        add_button.grid(row=0, column=0, padx=5, pady=0, sticky=W)

        update_button = Button(button_frame, text="Update", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.update_student)
        update_button.grid(row=0, column=1, padx=6, pady=0, sticky=W)

        delete_button = Button(button_frame, text="Delete", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.delete_student)
        delete_button.grid(row=0, column=2, padx=6, pady=0, sticky=W)

        clear_button = Button(button_frame, text="Clear", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.clear_student_data)
        clear_button.grid(row=0, column=3, padx=6, pady=0, sticky=W)


        search_frame = Frame(self.frame2, bd=3, relief=RIDGE)
        search_frame.place(x=5, y=5, width=725, height=40)

        self.var_search_by = StringVar()
        self.var_search = StringVar()

        search_by = Label(search_frame, text="Search By :", font=("verdana", 12, "bold"), width=10, bg="black", fg="white",bd=3,relief=RIDGE)
        search_by.grid(row=0, column=0, padx=1, pady=1, sticky=W)

        search_by_entry = ttk.Combobox(search_frame, textvariable=self.var_search_by, font=("verdana", 12, "bold"), width=12, state="readonly")
        search_by_entry["value"] = ("Student ID", "Student Name", "Contact No")
        search_by_entry.grid(row=0, column=1, padx=1, pady=1, sticky=W)

        search_field = ttk.Entry(search_frame, textvariable=self.var_search, font=("verdana", 12, "bold"), width=14)
        search_field.grid(row=0, column=2, padx=6, pady=1, sticky=W)

        search_button = Button(search_frame, text="Search", font=("verdana", 13, "bold"), width=10, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.search_student)
        search_button.grid(row=0, column=3, padx=0, pady=0, sticky=W)

        show_all_button = Button(search_frame, text="Show All", font=("verdana", 13, "bold"), width=10, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.load_student_data)
        show_all_button.grid(row=0, column=4, padx=9, pady=0, sticky=W)


        table_frame=Frame(self.frame2, bd=3, relief=RIDGE)
        table_frame.place(x=5, y=50, width=725, height=325)

        scroll_x = ttk.Scrollbar(table_frame, orient=HORIZONTAL)
        scroll_y = ttk.Scrollbar(table_frame, orient=VERTICAL)

        self.student_table = ttk.Treeview(table_frame, xscrollcommand=scroll_x.set,yscrollcommand=scroll_y.set, show="headings")

        scroll_x.pack(side=BOTTOM, fill=X)
        scroll_y.pack(side=RIGHT, fill=Y)

        self.student_table.pack(fill=BOTH, expand=1)

        scroll_x.config(command=self.student_table.xview)
        scroll_y.config(command=self.student_table.yview)

        self.student_table.bind("<<TreeviewSelect>>", self.load_data_to_entry)

        self.load_student_data()


        detail_frame = Frame(self.frame3, bd=3, relief=RIDGE)
        detail_frame.place(x=5, y=5, width=725, height=250)

        option_frame = Frame(self.frame3, bd=3, relief=RIDGE)
        option_frame.place(x=5, y=270, width=725, height=45)

        teacher_id = Label(detail_frame, text="Teacher ID:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        teacher_id.grid(row=0, column=0, padx=5, pady=3, sticky=W)

        teacher_id_entry = ttk.Entry(detail_frame, textvariable=self.id_var, font=("times new roman", 14, "bold"), width=19)
        teacher_id_entry.grid(row=0, column=1, padx=1, pady=3, sticky=W)

        teacher_name = Label(detail_frame, text="Teacher Name:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        teacher_name.grid(row=1, column=0, padx=5, pady=3, sticky=W)

        teacher_name_entry = ttk.Entry(detail_frame, textvariable=self.name_var, font=("times new roman", 14, "bold"), width=19)
        teacher_name_entry.grid(row=1, column=1, padx=1, pady=3, sticky=W)

        t_gender = Label(detail_frame, text="Gender:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        t_gender.grid(row=2, column=0, padx=5, pady=3, sticky=W)

        t_gender_entry = ttk.Combobox(detail_frame, textvariable=self.gender_var, font=("times new roman", 14, "bold"), width=17, state="readonly")
        t_gender_entry["value"] = ("Male", "Female", "Others")
        t_gender_entry.grid(row=2, column=1, padx=1, pady=3, sticky=W)

        t_email = Label(detail_frame, text="Email:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        t_email.grid(row=3, column=0, padx=5, pady=3, sticky=W)

        t_email_entry = ttk.Entry(detail_frame, textvariable=self.email_var, font=("times new roman", 14, "bold"), width=19)
        t_email_entry.grid(row=3, column=1, padx=1, pady=3, sticky=W)

        t_contact = Label(detail_frame, text="Contact No:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        t_contact.grid(row=4, column=0, padx=5, pady=3, sticky=W)

        t_contact_entry = ttk.Entry(detail_frame, textvariable=self.contact_var, font=("times new roman", 14, "bold"), width=19)
        t_contact_entry.grid(row=4, column=1, padx=1, pady=3, sticky=W)

        t_address = Label(detail_frame, text="Address:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        t_address.grid(row=5, column=0, padx=5, pady=3, sticky=W)

        t_address_entry = ttk.Entry(detail_frame, textvariable=self.address_var, font=("times new roman", 14, "bold"), width=19)
        t_address_entry.grid(row=5, column=1, padx=1, pady=3, sticky=W)

        joining_date = Label(detail_frame, text="Joining Date:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        joining_date.grid(row=6, column=0, padx=5, pady=3, sticky=W)

        joining_date_entry = tkcalendar.DateEntry(detail_frame, textvariable=self.joining_var, font=("times new roman", 14, "bold"), width=17)
        joining_date_entry.grid(row=6, column=1, padx=1, pady=3, sticky=W)

        t_department = Label(detail_frame, text="Department:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        t_department.grid(row=0, column=2, padx=18, pady=3, sticky=W)

        t_department_entry = ttk.Combobox(detail_frame, textvariable=self.dept_var, font=("times new roman", 14, "bold"), width=17, state="readonly")
        t_department_entry["value"] = ("CSE", "EEE", "ME", "CE", "IPE", "BME", "ChE", "WRE", "NAME", "URP", "ARCHI")
        t_department_entry.grid(row=0, column=3, padx=0, pady=3, sticky=W)

        current_position = Label(detail_frame, text="Current Position:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        current_position.grid(row=1, column=2, padx=18, pady=3, sticky=W)

        current_position_entry = ttk.Combobox(detail_frame, textvariable=self.position_var, font=("times new roman", 14, "bold"), width=17, state="readonly")
        current_position_entry["value"] = ("Lecturer", "Associate Professor", "Assistant Professor", "Professor")
        current_position_entry.grid(row=1, column=3, padx=0, pady=3, sticky=W)

        t_session = Label(detail_frame, text="Session:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        t_session.grid(row=2, column=2, padx=18, pady=3, sticky=W)

        t_session_entry = ttk.Combobox(detail_frame, textvariable=self.session_var, font=("times new roman", 14, "bold"), width=17, state="readonly")
        t_session_entry["value"] = ("2018-19", "2019-20", "2020-21","2021-22","2022-23","2023-24")
        t_session_entry.grid(row=2, column=3, padx=0, pady=3, sticky=W)

        courses_taken = Label(detail_frame, text="Courses Taken:", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        courses_taken.grid(row=3, column=2, padx=18, pady=3, sticky=W)

        courses_taken_entry = ttk.Entry(detail_frame, textvariable=self.courses_var, font=("times new roman", 14, "bold"), width=19)
        courses_taken_entry.grid(row=3, column=3, padx=0, pady=3, sticky=W)

        basic_salary = Label(detail_frame, text="Basic Salary ($):", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        basic_salary.grid(row=4, column=2, padx=18, pady=3, sticky=W)

        basic_salary_entry = ttk.Entry(detail_frame, textvariable=self.basic_var, font=("times new roman", 14, "bold"), width=19)
        basic_salary_entry.grid(row=4, column=3, padx=0, pady=3, sticky=W)

        other_salary = Label(detail_frame, text="Other Salary ($):", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        other_salary.grid(row=5, column=2, padx=18, pady=3, sticky=W)

        other_salary_entry = ttk.Entry(detail_frame, textvariable=self.other_var, font=("times new roman", 14, "bold"), width=19)
        other_salary_entry.grid(row=5, column=3, padx=0, pady=3, sticky=W)

        net_salary = Label(detail_frame, text="Net Salary ($):", font=("times new roman", 14, "bold"), bg="white", bd=0, relief=RIDGE)
        net_salary.grid(row=6, column=2, padx=18, pady=3, sticky=W)

        net_salary_entry = ttk.Entry(detail_frame, textvariable=self.net_var, font=("times new roman", 14, "bold"), width=19)
        net_salary_entry.grid(row=6, column=3, padx=0, pady=3, sticky=W)

        add_button = Button(option_frame, text="Add", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.add_teacher)
        add_button.grid(row=0, column=0, padx=5, pady=0, sticky=W)

        update_button = Button(option_frame, text="Update", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.update_teacher)
        update_button.grid(row=0, column=1, padx=6, pady=0, sticky=W)

        delete_button = Button(option_frame, text="Delete", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.delete_teacher)
        delete_button.grid(row=0, column=2, padx=6, pady=0, sticky=W)

        clear_button = Button(option_frame, text="Clear", font=("verdana", 14, "bold"), width=12, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.clear_teacher_data)
        clear_button.grid(row=0, column=3, padx=6, pady=0, sticky=W)


        search_frame2 = Frame(self.frame4, bd=3, relief=RIDGE)
        search_frame2.place(x=5, y=5, width=725, height=40)

        self.search_criteria_var = StringVar()
        self.search_value_var = StringVar()

        search_criteria = Label(search_frame2, text="Search By :", font=("verdana", 12, "bold"), width=10, bg="black", fg="white", bd=3, relief=RIDGE)
        search_criteria.grid(row=0, column=0, padx=1, pady=1, sticky=W)

        search_criteria_entry = ttk.Combobox(search_frame2, textvariable=self.search_criteria_var, font=("verdana", 12, "bold"), width=12, state="readonly")
        search_criteria_entry["value"] = ("Teacher ID", "Teacher Name", "Contact No", "Department")
        search_criteria_entry.grid(row=0, column=1, padx=1, pady=1, sticky=W)

        search_value = ttk.Entry(search_frame2, textvariable=self.search_value_var, font=("verdana", 12, "bold"), width=14)
        search_value.grid(row=0, column=2, padx=6, pady=1, sticky=W)

        t_search_button = Button(search_frame2, text="Search", font=("verdana", 13, "bold"), width=10, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.search_teacher)
        t_search_button.grid(row=0, column=3, padx=0, pady=0, sticky=W)

        t_show_all_button = Button(search_frame2, text="Show All", font=("verdana", 13, "bold"), width=10, bg="blue", fg="white", bd=3, relief=RIDGE, command=self.load_teacher_data)
        t_show_all_button.grid(row=0, column=4, padx=9, pady=0, sticky=W)

        table_frame2 = Frame(self.frame4, bd=3, relief=RIDGE)
        table_frame2.place(x=5, y=50, width=725, height=270)

        t_scroll_x = ttk.Scrollbar(table_frame2, orient=HORIZONTAL)
        t_scroll_y = ttk.Scrollbar(table_frame2, orient=VERTICAL)

        self.teacher_table = ttk.Treeview(table_frame2, xscrollcommand=t_scroll_x.set, yscrollcommand=t_scroll_y.set, show="headings")

        t_scroll_x.pack(side=BOTTOM, fill=X)
        t_scroll_y.pack(side=RIGHT, fill=Y)

        self.teacher_table.pack(fill=BOTH, expand=1)

        t_scroll_x.config(command=self.teacher_table.xview)
        t_scroll_y.config(command=self.teacher_table.yview)

        self.teacher_table.bind("<<TreeviewSelect>>", self.load_teacher_data_from_table)

        self.load_teacher_data()



    def load_student_data(self):
        self.var_id.set("")
        self.var_name.set("")
        self.var_gender.set("")
        self.var_email.set("")
        self.var_contact.set("")
        self.var_address.set("")
        self.var_dept.set("")
        self.var_session.set("")
        self.var_lt.set("")
        self.var_courses.set("")

        self.student_table.delete(*self.student_table.get_children())

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Student Form"]

            header_row = ws[1]
            columns = [col.value for col in header_row]

            self.student_table['columns'] = columns

            for col_name in columns:
                self.student_table.column(col_name, width=140, stretch=False)

            for col_name in columns:
                self.student_table.heading(col_name, text=col_name)

            for row in ws.iter_rows(min_row=2, values_only=True):
                self.student_table.insert("", "end", values=row)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while loading student data: {e}")

    def add_student(self, event=None):
        student_id = self.var_id.get()
        student_name = self.var_name.get()
        gender = self.var_gender.get()
        email = self.var_email.get()
        contact = self.var_contact.get()
        address = self.var_address.get()
        department = self.var_dept.get()
        session = self.var_session.get()
        level_term = self.var_lt.get()
        courses = self.var_courses.get()

        try:
            student_id = int(student_id)
        except ValueError:
            messagebox.showerror("Error", "Student ID must be an integer.")
            return

        if not (student_id and student_name and contact and department):
            messagebox.showerror("Error", "Please fill in all required fields correctly.")
            return

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Student Form"]

            exists = False
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if student_id in row:
                    exists = True
                    break

            if not exists:
                ws.append([student_id, student_name, gender, email, contact, address, department, session, level_term, courses])
                wb.save("School Management File.xlsx")
                messagebox.showinfo("Success", "Student data added successfully.")

            else:
                messagebox.showerror("Error", "Student with this ID already exists.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while adding student data: {e}")

        self.load_student_data()

    def update_student(self):
        student_id = self.var_id.get()
        student_name = self.var_name.get()
        gender = self.var_gender.get()
        email = self.var_email.get()
        contact = self.var_contact.get()
        address = self.var_address.get()
        department = self.var_dept.get()
        session = self.var_session.get()
        level_term = self.var_lt.get()
        courses = self.var_courses.get()

        try:
            student_id = int(student_id)
        except ValueError:
            messagebox.showerror("Error", "Student ID must be an integer.")
            return

        if not (student_id and student_name and contact and department):
            messagebox.showerror("Error", "Please fill in all required fields correctly.")
            return

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Student Form"]

            row_update = None
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if student_id in row:
                    row_update = row
                    break

            if row_update:
                for idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if student_id == row_data[0]:
                        row_index = idx
                        break

                ws.cell(row=row_index, column=2).value = student_name
                ws.cell(row=row_index, column=3).value = gender
                ws.cell(row=row_index, column=4).value = email
                ws.cell(row=row_index, column=5).value = contact
                ws.cell(row=row_index, column=6).value = address
                ws.cell(row=row_index, column=7).value = department
                ws.cell(row=row_index, column=8).value = session
                ws.cell(row=row_index, column=9).value = level_term
                ws.cell(row=row_index, column=10).value = courses

                wb.save("School Management File.xlsx")
                messagebox.showinfo("Success", "Student data updated successfully.")
            else:
                messagebox.showerror("Error", "Student with this ID does not exist.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while updating student data: {e}")

        self.load_student_data()

    def delete_student(self):
        selected_row = self.student_table.focus()
        student_id = self.student_table.item(selected_row)['values'][0]

        confirmation = messagebox.askyesno("Confirmation",
                                           "Are you sure you want to delete the selected teacher's data?")
        if confirmation:
            try:
                wb = openpyxl.load_workbook("School Management File.xlsx")
                ws = wb["Student Form"]

                for idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if student_id == row_data[0]:
                        row_index = idx
                        break

                ws.delete_rows(row_index)
                wb.save("School Management File.xlsx")

                messagebox.showinfo("Success", "Student data deleted successfully.")

                self.load_student_data()

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while deleting student data: {e}")

    def clear_student_data(self, event=None):
        self.var_id.set("")
        self.var_name.set("")
        self.var_gender.set("")
        self.var_email.set("")
        self.var_contact.set("")
        self.var_address.set("")
        self.var_dept.set("")
        self.var_session.set("")
        self.var_lt.set("")
        self.var_courses.set("")

    def load_data_to_entry(self, event=None):
        selected_row = self.student_table.focus()
        data_entry = self.student_table.item(selected_row)
        data_value = data_entry["values"]

        self.var_id.set(data_value[0])
        self.var_name.set(data_value[1])
        self.var_gender.set(data_value[2])
        self.var_email.set(data_value[3])
        self.var_contact.set(data_value[4])
        self.var_address.set(data_value[5])
        self.var_dept.set(data_value[6])
        self.var_session.set(data_value[7])
        self.var_lt.set(data_value[8])
        self.var_courses.set(data_value[9])

    def search_student(self):
        search_criteria = self.var_search_by.get()
        search_value = self.var_search.get()

        if not search_criteria:
            messagebox.showerror("Error", "Please select a search criteria.")
            return

        if not search_value:
            messagebox.showerror("Error", "Please enter a search value.")
            return

        self.student_table.delete(*self.student_table.get_children())

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Student Form"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if search_criteria == "Student ID" and str(row[0]) == search_value:
                    self.student_table.insert("", "end", values=row)
                elif search_criteria == "Student Name" and search_value.lower() in row[1].lower():
                    self.student_table.insert("", "end", values=row)
                elif search_criteria == "Contact No" and str(row[4]) == search_value:
                    self.student_table.insert("", "end", values=row)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while searching student data: {e}")



    def load_teacher_data(self):
        self.id_var.set("")
        self.name_var.set("")
        self.gender_var.set("")
        self.email_var.set("")
        self.contact_var.set("")
        self.address_var.set("")
        self.dept_var.set("")
        self.position_var.set("")
        self.joining_var.set("")
        self.session_var.set("")
        self.courses_var.set("")
        self.basic_var.set("")
        self.other_var.set("")
        self.net_var.set("")

        self.teacher_table.delete(*self.teacher_table.get_children())

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Teacher Form"]

            header_row = ws[1]
            columns = [col.value for col in header_row]

            self.teacher_table['columns'] = columns

            for col_name in columns:
                self.teacher_table.column(col_name, width=140, stretch=False)

            for col_name in columns:
                self.teacher_table.heading(col_name, text=col_name)

            for row in ws.iter_rows(min_row=2, values_only=True):
                self.teacher_table.insert("", "end", values=row)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while loading teacher data: {e}")

    def add_teacher(self, event=None):
        teacher_id = self.id_var.get()
        teacher_name = self.name_var.get()
        teacher_gender = self.gender_var.get()
        teacher_email = self.email_var.get()
        teacher_contact = self.contact_var.get()
        teacher_address = self.address_var.get()
        teacher_department = self.dept_var.get()
        teacher_current_position = self.position_var.get()
        teacher_joining_date = self.joining_var.get()
        teacher_session = self.session_var.get()
        teacher_courses = self.courses_var.get()
        teacher_basic_salary = self.basic_var.get()
        teacher_other_salary = self.other_var.get()
        teacher_net_salary = self.net_var.get()

        try:
            teacher_id = int(teacher_id)
        except ValueError:
            messagebox.showerror("Error", "Teacher ID must be an integer.")
            return

        if not (teacher_id and teacher_name and teacher_contact and teacher_department):
            messagebox.showerror("Error", "Please fill in all required fields correctly.")
            return

        try:
            teacher_basic_salary = float(teacher_basic_salary)
            teacher_other_salary = float(teacher_other_salary)
            if not teacher_net_salary:
                # Calculate net salary if not provided
                teacher_net_salary = teacher_basic_salary + teacher_other_salary
        except ValueError:
            messagebox.showerror("Error", "Salary fields must contain valid numeric values.")
            return

            # Check if the net salary matches the sum of basic salary and other salary
        if teacher_basic_salary + teacher_other_salary != teacher_net_salary:
            messagebox.showwarning("Warning", "Net salary does not match the sum of basic salary and other salary.")
            return

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Teacher Form"]

            exists = False
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if teacher_id in row:
                    exists = True
                    break

            if not exists:
                ws.append([teacher_id, teacher_name, teacher_gender, teacher_email, teacher_contact, teacher_address, teacher_joining_date, teacher_department,
                           teacher_current_position, teacher_session, teacher_courses, teacher_basic_salary, teacher_other_salary, teacher_net_salary])
                wb.save("School Management File.xlsx")
                messagebox.showinfo("Success", "Teacher data added successfully.")

            else:
                messagebox.showerror("Error", "Teacher with this ID already exists.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while adding student data: {e}")

        self.load_teacher_data()

    def update_teacher(self):
        teacher_id = self.id_var.get()
        teacher_name = self.name_var.get()
        teacher_gender = self.gender_var.get()
        teacher_email = self.email_var.get()
        teacher_contact = self.contact_var.get()
        teacher_address = self.address_var.get()
        teacher_department = self.dept_var.get()
        teacher_current_position = self.position_var.get()
        teacher_joining_date = self.joining_var.get()
        teacher_session = self.session_var.get()
        teacher_courses = self.courses_var.get()
        teacher_basic_salary = self.basic_var.get()
        teacher_other_salary = self.other_var.get()
        teacher_net_salary = self.net_var.get()

        try:
            teacher_id = int(teacher_id)
        except ValueError:
            messagebox.showerror("Error", "Teacher ID must be an integer.")
            return

        if not (teacher_id and teacher_name and teacher_contact and teacher_department):
            messagebox.showerror("Error", "Please fill in all required fields correctly.")
            return

        try:
            teacher_basic_salary = float(teacher_basic_salary)
            teacher_other_salary = float(teacher_other_salary)
            if not teacher_net_salary:
                # Calculate net salary if not provided
                teacher_net_salary = teacher_basic_salary + teacher_other_salary
        except ValueError:
            messagebox.showerror("Error", "Salary fields must contain valid numeric values.")
            return

            # Check if the net salary matches the sum of basic salary and other salary
        if teacher_basic_salary + teacher_other_salary != teacher_net_salary:
            messagebox.showwarning("Warning", "Net salary does not match the sum of basic salary and other salary.")
            return

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Teacher Form"]

            row_update = None
            for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
                if teacher_id in row:
                    row_update = row
                    break

            if row_update:
                for idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if teacher_id == row_data[0]:
                        row_index = idx
                        break

                ws.cell(row=row_index, column=2).value = teacher_name
                ws.cell(row=row_index, column=3).value = teacher_gender
                ws.cell(row=row_index, column=4).value = teacher_email
                ws.cell(row=row_index, column=5).value = teacher_contact
                ws.cell(row=row_index, column=6).value = teacher_address
                ws.cell(row=row_index, column=7).value = teacher_joining_date
                ws.cell(row=row_index, column=8).value = teacher_department
                ws.cell(row=row_index, column=9).value = teacher_current_position
                ws.cell(row=row_index, column=10).value = teacher_session
                ws.cell(row=row_index, column=11).value = teacher_courses
                ws.cell(row=row_index, column=12).value = teacher_basic_salary
                ws.cell(row=row_index, column=13).value = teacher_other_salary
                ws.cell(row=row_index, column=14).value = teacher_net_salary

                wb.save("School Management File.xlsx")
                messagebox.showinfo("Success", "Teacher data updated successfully.")
            else:
                messagebox.showerror("Error", "Teacher with this ID does not exist. Add teacher instead.")

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while updating student data: {e}")

        self.load_teacher_data()

    def delete_teacher(self):
        selected_row = self.teacher_table.focus()
        teacher_id = self.teacher_table.item(selected_row)['values'][0]

        confirmation = messagebox.askyesno("Confirmation",
                                           "Are you sure you want to delete the selected teacher's data?")
        if confirmation:
            try:
                wb = openpyxl.load_workbook("School Management File.xlsx")
                ws = wb["Teacher Form"]

                for idx, row_data in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if teacher_id == row_data[0]:
                        row_index = idx
                        break

                ws.delete_rows(row_index)
                wb.save("School Management File.xlsx")

                messagebox.showinfo("Success", "Teacher data deleted successfully.")

                self.load_teacher_data()

            except Exception as e:
                messagebox.showerror("Error", f"An error occurred while deleting teacher data: {e}")

    def clear_teacher_data(self, event=None):
        self.id_var.set("")
        self.name_var.set("")
        self.gender_var.set("")
        self.email_var.set("")
        self.contact_var.set("")
        self.address_var.set("")
        self.dept_var.set("")
        self.position_var.set("")
        self.joining_var.set("")
        self.session_var.set("")
        self.courses_var.set("")
        self.basic_var.set("")
        self.other_var.set("")
        self.net_var.set("")

    def load_teacher_data_from_table(self, event=None):
        selected_row = self.teacher_table.focus()
        data_entry = self.teacher_table.item(selected_row)
        data_value = data_entry["values"]

        self.id_var.set(data_value[0])
        self.name_var.set(data_value[1])
        self.gender_var.set(data_value[2])
        self.email_var.set(data_value[3])
        self.contact_var.set(data_value[4])
        self.address_var.set(data_value[5])
        self.joining_var.set(data_value[6])
        self.dept_var.set(data_value[7])
        self.position_var.set(data_value[8])
        self.session_var.set(data_value[9])
        self.courses_var.set(data_value[10])
        self.basic_var.set(data_value[11])
        self.other_var.set(data_value[12])
        self.net_var.set(data_value[13])

    def search_teacher(self):
        search_criteria = self.search_criteria_var.get()
        search_value = self.search_value_var.get()

        if not search_criteria:
            messagebox.showerror("Error", "Please select a search criteria.")
            return

        if not search_value:
            messagebox.showerror("Error", "Please enter a search value.")
            return

        self.teacher_table.delete(*self.teacher_table.get_children())

        try:
            wb = openpyxl.load_workbook("School Management File.xlsx")
            ws = wb["Teacher Form"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if search_criteria == "Teacher ID" and str(row[0]) == search_value:
                    self.teacher_table.insert("", "end", values=row)
                elif search_criteria == "Teacher Name" and search_value.lower() in row[1].lower():
                    self.teacher_table.insert("", "end", values=row)
                elif search_criteria == "Contact No" and str(row[4]) == search_value:
                    self.teacher_table.insert("", "end", values=row)
                elif search_criteria == "Department" and search_value.lower() in row[7].lower():
                    self.teacher_table.insert("", "end", values=row)

        except Exception as e:
            messagebox.showerror("Error", f"An error occurred while searching student data: {e}")


root = Tk()
app = Student(root)
root.mainloop()